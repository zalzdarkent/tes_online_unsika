import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Template Bank Soal'

# Define headers (removed 'tags')
headers = [
    'title',
    'jenis_soal',
    'pertanyaan',
    'opsi_a',
    'opsi_b',
    'opsi_c',
    'opsi_d',
    'jawaban_benar',
    'skor',
    'difficulty_level',
    'is_public',
    'kategori_tes_id',
    'equation',
    'skala_min',
    'skala_maks',
    'skala_label_min',
    'skala_label_maks'
]

# Add headers
ws.append(headers)

# Style for header
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply header styling
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Add example data (removed tags column)
examples = [
    [
        'Soal Pilihan Ganda - Geografi Indonesia',
        'pilihan_ganda',
        'Apa ibu kota Indonesia?',
        'Jakarta',
        'Bandung',
        'Surabaya',
        'Medan',
        'a',
        10,
        'easy',
        'true',
        1,
        '',
        '',
        '',
        '',
        ''
    ],
    [
        'Soal Multi Choice - Bahasa Pemrograman',
        'multi_choice',
        'Pilih bahasa pemrograman berikut (pilih semua yang benar)',
        'Python',
        'Java',
        'HTML',
        'CSS',
        'a,b',
        15,
        'medium',
        'true',
        1,
        '',
        '',
        '',
        '',
        ''
    ],
    [
        'Soal Esai - Teori Programming',
        'esai',
        'Jelaskan perbedaan antara OOP dan Procedural Programming',
        '',
        '',
        '',
        '',
        'Jawaban bervariasi',
        20,
        'hard',
        'false',
        2,
        '',
        '',
        '',
        '',
        ''
    ],
    [
        'Soal Skala - Survey Kepuasan',
        'skala',
        'Seberapa puas Anda dengan layanan kami?',
        '',
        '',
        '',
        '',
        '5',
        5,
        'easy',
        'true',
        '',
        '',
        1,
        5,
        'Sangat Tidak Puas',
        'Sangat Puas'
    ]
]

# Add examples to worksheet
for example in examples:
    ws.append(example)

# Apply border and alignment to all data cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# Set column widths (updated for removed tags column)
col_widths = [35, 20, 50, 25, 25, 25, 25, 20, 10, 18, 12, 18, 15, 12, 12, 22, 22]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Set row heights
ws.row_dimensions[1].height = 30
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 40

# Save workbook
wb.save('public/template-bank-soal.xlsx')
print('✅ File Excel template-bank-soal.xlsx berhasil dibuat!')
